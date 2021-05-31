# To convert txt file to xlsx file.

import openpyxl
import codecs
from openpyxl.utils import get_column_letter

def txt_to_xlsx(filename, outfile):
    fr = codecs.open(filename, 'r')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws = wb.create_sheet()
    ws.title = 'Sheet1'
    row = 0
    for line in fr:
        row += 1
        line = line.strip()
        # line = line.split('\t')
        if ',' in line:
            line = line.split(',')
            # print(line)
        else:
            line = line.split('\t')
        col = 0
        # print(len(line))
        for j in range(len(line)):
            col += 1
            # print (line[j])
            ws.cell(column=col, row=row, value=line[j].format(get_column_letter(col)))
    wb.save(outfile)
